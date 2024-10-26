import copy
import openpyxl
from openpyxl.utils import get_column_letter


def getFromExcel(fileName):
    # 从第二行开始获取关键字的子选项,返回字典
    workBook = openpyxl.load_workbook(fileName)
    workSheet = workBook.active

    contentFileDict = {}
    for row in range(2, workSheet.max_row + 1):
        if workSheet['A'+str(row)].value not in contentFileDict:
            contentFileDict[workSheet['A'+str(row)].value] = [workSheet['B'+str(row)].value]
        else:
            contentFileDict[workSheet['A'+str(row)].value].append(workSheet['B'+str(row)].value)

    workBook.close()
    return contentFileDict

def getDetail(fileName, keyDict):
    """
    detailDict = { DAPING: {DAPING_XH:.., DAPING_DJ:..,},
                   CHULIQI:{CHULIQI_XH:.., CHULIQI_DJ:..,},
                   ...
                 }
    """
    workBook = openpyxl.load_workbook(fileName)
    workSheet = workBook.active

    detailDict = {}
    for row in range(2, workSheet.max_row + 1):
        if workSheet['B'+str(row)].value in keyDict.values():
            detailDict[workSheet['B'+str(row)].value] = {}
            for column in range(3,workSheet.max_column+1):
                detailDict[workSheet['B'+str(row)].value][workSheet['A'+str(row)].value+'_'+workSheet[chr(column+64)+'1'].value] = workSheet[chr(column+64)+str(row)].value

    # print(detailDict)

    workBook.close()
    return detailDict

def copyExcel(originFileName, newFileName, replaceDict):
    path = originFileName
    save_path = newFileName

    wb = openpyxl.load_workbook(path)
    wb2 = openpyxl.Workbook()

    sheetnames = wb.sheetnames
    for sheetname in sheetnames:
        # print(sheetname)
        sheet = wb[sheetname]
        sheet2 = wb2.create_sheet(sheetname)

        # tab颜色
        sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

        # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
        wm = list(sheet.merged_cells)
        if len(wm) > 0:
            for i in range(0, len(wm)):
                cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
                sheet2.merge_cells(cell2)

        for i, row in enumerate(sheet.iter_rows()):
            sheet2.row_dimensions[i+1].height = sheet.row_dimensions[i+1].height
            for j, cell in enumerate(row):
                sheet2.column_dimensions[get_column_letter(j+1)].width = sheet.column_dimensions[get_column_letter(j+1)].width

                # 查找是否有关键字,并替换内容
                if cell.value in replaceDict.keys():
                    sheet2.cell(row=i + 1, column=j + 1, value=replaceDict[cell.value])
                else:
                    sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

                # 设置单元格格式
                source_cell = sheet.cell(i+1, j+1)
                target_cell = sheet2.cell(i+1, j+1)
                target_cell.fill = copy.copy(source_cell.fill)
                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = copy.copy(source_cell.number_format)
                    target_cell.protection = copy.copy(source_cell.protection)
                    target_cell.alignment = copy.copy(source_cell.alignment)

    if 'Sheet' in wb2.sheetnames:
        del wb2['Sheet']
    wb2.save(save_path)

    wb.close()
    wb2.close()


if __name__ == "__main__":
    d = getFromExcel('data.xlsx')
    print(d)
